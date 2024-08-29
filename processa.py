from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from dotenv import load_dotenv
import urllib.request
import json
import os

def captura_dados(local,data_inicio,data_fim):
    load_dotenv()
    id = os.getenv('ID')
    try:
        url = (f"https://weather.visualcrossing.com/VisualCrossingWebServices/rest/services/timeline/{local}/"
               f"{data_inicio}/{data_fim}"
               f"?unitGroup=metric&include=statsfcst%2Cfcst%2Chours&key={id}&lang=pt&contentType=json")
        response = urllib.request.urlopen(url)
        data = json.load(response)



        return data
    except urllib.error.HTTPError as e:
        error_info = e.read().decode()
        print('Error code:', e.code, error_info)
        return None
    except urllib.error.URLError as e:
        error_info = e.read().decode()
        print('Error code:', e.code, error_info)
        return None

def save_excel(weather_data,file_name):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Previsão do Tempo'
    headers = ['Local','Data','Temp (C)', 'Temp Max(C)','Temp Min(C)',
               'Condição','Precipitação (mm)','Prob Precipitação(%)','Cobert Precipitação','Horas Precipitação']
    cor = PatternFill(start_color="4682B4", end_color="4682B4", fill_type="solid")
    fonte_cor = Font(color="FFFFFF")

    localizacao = weather_data.get("address")
    for col_index, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_index, value=header)
        cell.fill = cor
        cell.font = fonte_cor
    if weather_data and "days" in weather_data:
        row = 2

        for day in weather_data["days"]:
            date = day["datetime"]
            temp = day.get("temp", "N/A")
            temp_max = day.get("tempmax", "N/A")
            temp_min = day.get("tempmin", "N/A")
            condition = day.get("conditions", "N/A")
            precip = day.get("precip", "N/A")
            precip_prob = day.get("precipprob", "N/A")
            precip_cover = day.get("precipcover", "N/A")
            hours_precip = (precip_cover * 24) / 100


            ws.cell(row=row, column=1, value=localizacao)
            ws.cell(row=row, column=2, value=date)
            ws.cell(row=row, column=3, value=temp)
            ws.cell(row=row, column=4, value=temp_max)
            ws.cell(row=row, column=5, value=temp_min)
            ws.cell(row=row, column=6, value=condition)
            ws.cell(row=row, column=7, value=precip)
            ws.cell(row=row, column=8, value=precip_prob)
            ws.cell(row=row, column=9, value=precip_cover)
            ws.cell(row=row, column=10, value=hours_precip)

            precip_hours = ws.cell(row=row, column=10, value=hours_precip)
            precip_hours.number_format = '[h]:mm:ss'
            ws.cell(row=row, column=10,value=hours_precip/24)

            row += 1

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
    wb.save(file_name)
